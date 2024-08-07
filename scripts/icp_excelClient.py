import pandas as pd
from openpyxl import Workbook
from dotenv import load_dotenv
import numpy as np
import datetime as dt
import os as OS

load_dotenv()
master_path = OS.getenv("ICP_MASTER")
lista_nombres = OS.getenv("LISTA_NOMBRES")

OG_EXCEL = pd.read_excel(master_path)
excel = OG_EXCEL

excel = OG_EXCEL[OG_EXCEL["FECHA DE FACTURA"].str.contains(r"\b1\b")]


# kitExcel = excel[["NUMERO DE KIT", "REMPLAZO KIT"]]

# kitExcel.loc["NUMERO DE KIT"] = kitExcel.apply(
#     lambda x: x["REMPLAZO KIT"] if pd.notna(x["REMPLAZO KIT"]) else x["NUMERO DE KIT"],
#     axis=1,
# )


header_datetime = dt.datetime(2024, 7, 1, 0, 0)
columnData = {}
if header_datetime in excel.columns:
    columnData = excel[header_datetime]


clientName = excel["CLIENTE"].to_dict()
orderPurchase = excel["OC CLIENTE"].to_dict()
kit = excel["NUMERO DE KIT"].to_dict()
ubicacion = excel["UBICACIÓN"].to_dict()
numeroDeServicios = excel["# DE SERVICIOS"].to_dict()
wb = Workbook()
print(clientName)
ws = wb.active
ws.title = "Lista de Nombres"

ws["A1"] = "Nombre"
ws["B1"] = "Folio"
ws["C1"] = "Oc"
ws["D1"] = "KIT"
ws["E1"] = "Ubicacion"
ws["F1"] = "Sercicios"

for idx, item in enumerate(clientName.values(), start=2):
    text = item
    ws.cell(row=idx, column=1, value=text)

for idx, item in enumerate(columnData, start=2):
    text = item
    ws.cell(row=idx, column=2, value=text)

for idx, item in enumerate(orderPurchase.values(), start=2):
    text = item
    ws.cell(row=idx, column=3, value=text)

for idx, item in enumerate(kit.values(), start=2):
    text = item
    ws.cell(row=idx, column=4, value=text)

for idx, item in enumerate(ubicacion.values(), start=2):
    text = item
    ws.cell(row=idx, column=5, value=text)

for idx, item in enumerate(numeroDeServicios.values(), start=2):
    text = item
    ws.cell(row=idx, column=6, value=text)

wb.save(lista_nombres)
